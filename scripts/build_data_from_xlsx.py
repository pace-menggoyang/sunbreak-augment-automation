"""Regenerate data/*.json from the datamined "Max Possible Qurio Aug" xlsx.

Run this whenever the source spreadsheet is updated to a newer game
version. Nothing at runtime depends on openpyxl or the xlsx itself --
only the generated JSON files under data/ are used by the automation.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "Max Possible Qurio Aug (16.0.0).xlsx"
DATA_DIR = ROOT / "data"


def build_armor_pools(wb: openpyxl.Workbook) -> list[dict]:
    """ArmorAugPool sheet: armor series -> aug pool id + budget."""
    ws = wb["ArmorAugPool"]
    armors = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, armor_id, pool, budget = row[0], row[1], row[2], row[3]
        if name is None or armor_id is None:
            continue
        armors.append(
            {
                "armor_series": str(name).strip(),
                "armor_id": int(armor_id),
                "aug_pool": int(pool),
                "budget": int(budget),
            }
        )
    return armors


def build_skills(wb: openpyxl.Workbook) -> list[dict]:
    """Cost-Skill sheet: repeated (Skill Id, Cost, Skill Name) column groups."""
    ws = wb["Cost-Skill"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    skills: dict[str, dict] = {}
    col = 0
    while col < len(header):
        if header[col] == "Skill Id":
            for row in ws.iter_rows(min_row=2, values_only=True):
                skill_id, cost, name = row[col], row[col + 1], row[col + 2]
                if skill_id is None or name is None:
                    continue
                clean_name = str(name).strip()
                # Keep the first (lowest) cost tier seen for a name -- the
                # sheet lists each skill under exactly one cost bucket.
                if clean_name not in skills:
                    skills[clean_name] = {
                        "skill_id": int(skill_id),
                        "cost": int(cost),
                        "name": clean_name,
                    }
            col += 4  # Skill Id, Cost, Skill Name, blank separator
        else:
            col += 1

    return sorted(skills.values(), key=lambda s: s["name"])


def build_aug_pool_values(wb: openpyxl.Workbook) -> dict:
    """AugPoolValue sheet: per-pool list of possible augment entries.

    Reference/tooling only (e.g. sanity-checking OCR output against what's
    theoretically possible for a pool) -- not required by the live
    read/decide/act loop, which acts on what's actually on screen.
    """
    ws = wb["AugPoolValue"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[1]

    # The sheet repeats an 11-column block (AugPool,id,Desc,Lv1,Lv2,Lv3,
    # Lv1%,Lv2%,Lv3%,Cost,blank) side by side per data-mine group; find
    # each block's start by locating "AugPool" in the header.
    block_starts = [i for i, v in enumerate(header) if v == "AugPool"]

    pools: dict[int, list[dict]] = {}
    for start in block_starts:
        for row in rows[2:]:
            pool = row[start]
            if pool is None or not isinstance(pool, (int, float)):
                continue
            desc = row[start + 2]
            if desc is None:
                continue
            entry = {
                "skill_or_effect_id": int(row[start + 1]),
                "desc": str(desc).strip(),
                "lv1": row[start + 3],
                "lv2": row[start + 4],
                "lv3": row[start + 5],
                "lv1_pct": row[start + 6],
                "lv2_pct": row[start + 7],
                "lv3_pct": row[start + 8],
                "cost": row[start + 9],
            }
            pools.setdefault(int(pool), []).append(entry)

    return {str(k): v for k, v in sorted(pools.items())}


def apply_skill_overrides(skills: list[dict]) -> list[dict]:
    """Merge in data/skills_overrides.json -- skills confirmed to exist
    in-game (e.g. spotted during calibration against a live roll) but
    missing from the datamined xlsx, presumably added in a game update
    after the sheet's version. Keeping these in a separate override file
    (rather than hand-editing skills.json) means they survive re-running
    this script when the xlsx is updated. cost/skill_id may be null if
    unknown -- decision.py only ever matches on name, so that's fine.
    """
    overrides_path = DATA_DIR / "skills_overrides.json"
    if not overrides_path.exists():
        return skills
    overrides = json.loads(overrides_path.read_text())
    existing_names = {s["name"] for s in skills}
    added = [
        {"skill_id": o["skill_id"], "name": o["name"], "cost": o["cost"]}
        for o in overrides
        if o["name"] not in existing_names
    ]  # "source" is documentation for the overrides file, not carried into skills.json
    if added:
        print(f"merging {len(added)} skill(s) from skills_overrides.json: "
              f"{', '.join(o['name'] for o in added)}")
    return sorted(skills + added, key=lambda s: s["name"])


def main() -> None:
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    DATA_DIR.mkdir(exist_ok=True)

    armors = build_armor_pools(wb)
    (DATA_DIR / "armor_aug_pools.json").write_text(
        json.dumps(armors, indent=2, ensure_ascii=False) + "\n"
    )
    print(f"wrote {len(armors)} armor entries -> data/armor_aug_pools.json")

    skills = apply_skill_overrides(build_skills(wb))
    (DATA_DIR / "skills.json").write_text(
        json.dumps(skills, indent=2, ensure_ascii=False) + "\n"
    )
    print(f"wrote {len(skills)} skills -> data/skills.json")

    pool_values = build_aug_pool_values(wb)
    (DATA_DIR / "aug_pool_values.json").write_text(
        json.dumps(pool_values, indent=2, ensure_ascii=False) + "\n"
    )
    print(f"wrote {len(pool_values)} pools -> data/aug_pool_values.json")


if __name__ == "__main__":
    main()
